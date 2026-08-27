"""
database_manager.py
-------------------
plaka_DATABASE.xlsx dosyasını PROGRAM BAŞINDA BİR KEZ belleğe yükler.
Giriş/çıkış güncellemeleri in-memory yapılır; diske kayıt arka plan thread'inde çalışır.
Bu sayede hiçbir zaman ana thread / kamera thread GIL'i beklemez → sıfır FPS drop.

Primary Key : Plaka
Toggle mantığı :
  İçeride        → ÇIKIŞ YAPTI (çıkış saati yaz, süre hesapla)
  Çıkış Yapıldı  → GİRİŞ YAPTI (giriş saati yaz, çıkış temizle)
  K.Liste=True   → KARA LİSTE uyarısı
  Bulunamazsa    → YABANCI ARAÇ
"""

import json
import os
import threading
from datetime import datetime

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Sabitler ──────────────────────────────────────────────────────────────────
EXCEL_PATH    = "plaka_DATABASE.xlsx"
JSON_LOG_PATH = "vehicle_log.json"

DURUM_ICERIDE = "İçeride"
DURUM_CIKIS   = "Çıkış Yapıldı"

# Sütun indeksleri (0-tabanlı, veri satırında)
COL_PLAKA  = 1
COL_AMAC   = 2
COL_DURUM  = 3
COL_GIRIS  = 4
COL_CIKIS  = 5
COL_SURUCU = 6
COL_FIRMA  = 7
COL_KLIST  = 8

# ── In-memory workbook ────────────────────────────────────────────────────────
# Workbook bir kez yüklenir; tüm okuma/yazma RAM üzerinden yapılır.
# Diske kayıt arka plan thread'inde (asenkron) gerçekleşir.
_wb = None
_ws = None
_wb_lock   = threading.Lock()   # in-memory erişim kilidi
_save_lock = threading.Lock()   # eş zamanlı kayıt kilidi


def _init():
    """Modül yüklenirken workbook'u belleğe al."""
    global _wb, _ws
    if not OPENPYXL_AVAILABLE:
        return
    if not os.path.exists(EXCEL_PATH):
        print(f"[database_manager] UYARI: {EXCEL_PATH} bulunamadı.")
        return
    try:
        _wb = openpyxl.load_workbook(EXCEL_PATH)
        _ws = _wb.active
        print(f"[database_manager] Excel belleğe yüklendi: {_ws.max_row - 1} kayıt.")
    except Exception as e:
        print(f"[database_manager] Excel yükleme hatası: {e}")


_init()


def _save_async():
    """Workbook'u arka plan thread'inde güvenli ve atomik şekilde diske yazar."""
    def _do():
        with _save_lock:
            try:
                temp_path = EXCEL_PATH + ".tmp"
                _wb.save(temp_path)
                if os.path.exists(temp_path) and os.path.getsize(temp_path) > 1000:
                    if os.path.exists(EXCEL_PATH):
                        os.replace(temp_path, EXCEL_PATH)
                    else:
                        os.rename(temp_path, EXCEL_PATH)
            except Exception as e:
                print(f"[database_manager] Excel kayıt hatası: {e}")
    t = threading.Thread(target=_do, daemon=False)
    t.start()


# ── Yardımcılar ───────────────────────────────────────────────────────────────

def _safe(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    try:
        s = s.encode('cp1252').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    return s


import difflib

# Modelin sıklıkla karıştırabileceği görsel karakter çiftleri
CONFUSABLE_PAIRS = {
    ('O', 'D'), ('D', 'O'),
    ('O', '0'), ('0', 'O'),
    ('O', 'Q'), ('Q', 'O'),
    ('B', '8'), ('8', 'B'),
    ('I', '1'), ('1', 'I'),
    ('L', '1'), ('1', 'L'),
    ('Z', '2'), ('2', 'Z'),
    ('S', '5'), ('5', 'S'),
    ('G', '6'), ('6', 'G'),
    ('A', '4'), ('4', 'A')
}

def calculate_plate_similarity(p1: str, p2: str) -> float:
    """
    İki plaka arasındaki karakter ve optik benzerlik oranını hesaplar (0.0 - 1.0).
    """
    p1 = p1.upper().replace(' ', '')
    p2 = p2.upper().replace(' ', '')
    if p1 == p2:
        return 1.0
    
    max_len = max(len(p1), len(p2))
    if max_len == 0 or abs(len(p1) - len(p2)) > 1:
        return 0.0
        
    # Karakter bazlı ağırlıklı eşleşme
    if len(p1) == len(p2):
        score = 0.0
        for c1, c2 in zip(p1, p2):
            if c1 == c2:
                score += 1.0
            elif (c1, c2) in CONFUSABLE_PAIRS:
                score += 0.90 # Karıştırılabilir görsel benzerlik puanı (O/D, B/8, I/1)
        return score / max_len
    else:
        return difflib.SequenceMatcher(None, p1, p2).ratio()


def _find_row(plaka_clean: str):
    """
    Plakanın satır numarasını ve eşleşen veritabanı plakasını döner:
    1. Birebir tam eşleşme arar (%100).
    2. Tam eşleşme yoksa veritabanındaki kayıtlarla Bulanık Eşleştirme (Fuzzy Match) yapar.
       Eğer benzerlik >= %85 ise o satırı döner (Örn: 34 OPR 073 -> 34 DPR 073).
    Dönüş: (row_idx, matched_db_plaka, similarity_ratio) veya (None, None, 0.0)
    """
    if _ws is None:
        return None, None, 0.0
        
    # 1. Aşama: Birebir Tam Eşleşme
    for row_idx in range(2, _ws.max_row + 1):
        cell_val = _safe(_ws.cell(row=row_idx, column=COL_PLAKA + 1).value)
        db_p = cell_val.upper().replace(" ", "")
        if db_p == plaka_clean:
            return row_idx, db_p, 1.0
            
    # 2. Aşama: %85 ve Üzeri Bulanık Eşleştirme (Fuzzy Matching)
    best_row = None
    best_db_plaka = None
    best_similarity = 0.0
    
    for row_idx in range(2, _ws.max_row + 1):
        cell_val = _safe(_ws.cell(row=row_idx, column=COL_PLAKA + 1).value)
        db_p = cell_val.upper().replace(" ", "")
        if not db_p:
            continue
            
        sim = calculate_plate_similarity(plaka_clean, db_p)
        if sim > best_similarity:
            best_similarity = sim
            best_row = row_idx
            best_db_plaka = db_p
            
    if best_similarity >= 0.85 and best_row is not None:
        print(f"[Akilli Plaka Esleme] Okunan: {plaka_clean} -> Veritabanindaki: {best_db_plaka} (Eslesme: %{best_similarity*100:.1f})")
        return best_row, best_db_plaka, best_similarity
        
    return None, None, 0.0


def _append_json_log(entry: dict):
    """Hareketi vehicle_log.json'a (tarihe göre) ekler/günceller."""
    today = datetime.now().strftime("%Y-%m-%d")
    log = {}
    if os.path.exists(JSON_LOG_PATH):
        try:
            with open(JSON_LOG_PATH, "r", encoding="utf-8") as f:
                log = json.load(f)
        except Exception:
            log = {}

    if today not in log:
        log[today] = []

    plaka = entry.get("plaka", "")
    updated = False
    for i, rec in enumerate(log[today]):
        if rec.get("plaka") == plaka:
            log[today][i] = entry
            updated = True
            break
    if not updated:
        log[today].append(entry)

    try:
        with open(JSON_LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[database_manager] JSON log hatası: {e}")


# ── Ana fonksiyon ─────────────────────────────────────────────────────────────

def process_plate(plaka: str) -> dict:
    """
    Plakayı işler.

    Dönüş dict:
        status  : "giris" | "cikis" | "kara_liste" | "yabanci" | "hata"
        plaka, surucu, firma, amac, giris, cikis, sure, mesaj
    """
    now = datetime.now()
    plaka_clean = plaka.strip().upper().replace(" ", "")

    result = {
        "status": "hata",
        "plaka":  plaka_clean,
        "surucu": "",
        "firma":  "",
        "amac":   "",
        "giris":  "",
        "cikis":  "",
        "sure":   "",
        "mesaj":  "İşlem sırasında hata oluştu.",
    }

    if not OPENPYXL_AVAILABLE or _wb is None:
        result["mesaj"] = "Excel veritabanı yüklenemedi!"
        return result

    with _wb_lock:   # In-memory erişim — GIL yalnızca bu blok boyunca tutulur (hızlı)
        try:
            row_idx, matched_plaka, similarity = _find_row(plaka_clean)

            # ── YABANCI ARAÇ ─────────────────────────────────────────────────
            if row_idx is None:
                result["status"] = "yabanci"
                result["mesaj"]  = "YABANCI ARAÇ — Veritabanında kayıtlı değil"
                _append_json_log({
                    "plaka":  plaka_clean, "surucu": "", "firma": "", "amac": "",
                    "giris":  "", "cikis": "", "sure": "",
                    "durum":  "Yabancı Araç", "zaman": now.strftime("%H:%M:%S"),
                })
                return result

            # Eğer bulanık eşleşmeyle düzeltildiyse sonucu resmi veritabanı plakası yap
            if matched_plaka and matched_plaka != plaka_clean:
                result["plaka"] = matched_plaka

            # Satır verilerini oku
            def cell(col_idx):
                return _ws.cell(row=row_idx, column=col_idx + 1).value

            surucu   = _safe(cell(COL_SURUCU))
            firma    = _safe(cell(COL_FIRMA))
            amac     = _safe(cell(COL_AMAC))
            durum    = _safe(cell(COL_DURUM))
            klist    = cell(COL_KLIST)
            giris_dt = cell(COL_GIRIS)

            result["surucu"] = surucu
            result["firma"]  = firma
            result["amac"]   = amac

            # ── KARA LİSTE ──────────────────────────────────────────────────
            if str(klist).strip().lower() in ("true", "1", "yes", "evet"):
                result["status"] = "kara_liste"
                result["mesaj"]  = f"KARA LİSTE — {surucu} | {firma}"
                return result

            # ── GİRİŞ / ÇIKIŞ toggle ────────────────────────────────────────
            if durum == DURUM_ICERIDE:
                # → ÇIKIŞ YAPTI
                cikis_str = now.strftime("%H:%M:%S")
                sure_str  = ""
                if isinstance(giris_dt, datetime):
                    total_sec = int((now - giris_dt).total_seconds())
                    total_min = total_sec // 60
                    saat, dakika = divmod(total_min, 60)
                    saniye = total_sec % 60
                    if saat > 0:
                        sure_str = f"{saat}sa {dakika}dk"
                    elif dakika > 0:
                        sure_str = f"{dakika}dk {saniye}sn"
                    else:
                        sure_str = f"{saniye}sn"

                # In-memory güncelle
                _ws.cell(row=row_idx, column=COL_DURUM + 1).value = DURUM_CIKIS
                _ws.cell(row=row_idx, column=COL_CIKIS  + 1).value = now

                result.update({
                    "status": "cikis",
                    "giris":  giris_dt.strftime("%H:%M:%S") if isinstance(giris_dt, datetime) else "",
                    "cikis":  cikis_str,
                    "sure":   sure_str,
                    "mesaj":  f"ÇIKIŞ YAPTI | {surucu} | {firma} | {amac}"
                              + (f" | Tesis içi süre: {sure_str}" if sure_str else ""),
                })
                _append_json_log({
                    "plaka":  plaka_clean, "surucu": surucu, "firma": firma, "amac": amac,
                    "giris":  result["giris"], "cikis": cikis_str,
                    "sure":   sure_str, "durum": DURUM_CIKIS,
                })

            else:
                # → GİRİŞ YAPTI
                giris_str = now.strftime("%H:%M:%S")

                # In-memory güncelle
                _ws.cell(row=row_idx, column=COL_DURUM + 1).value = DURUM_ICERIDE
                _ws.cell(row=row_idx, column=COL_GIRIS  + 1).value = now
                _ws.cell(row=row_idx, column=COL_CIKIS  + 1).value = None

                result.update({
                    "status": "giris",
                    "giris":  giris_str,
                    "cikis":  "",
                    "mesaj":  f"GİRİŞ YAPTI | {surucu} | {firma} | {amac}",
                })
                _append_json_log({
                    "plaka":  plaka_clean, "surucu": surucu, "firma": firma, "amac": amac,
                    "giris":  giris_str, "cikis": "", "sure": "", "durum": DURUM_ICERIDE,
                })

        except Exception as e:
            result["status"] = "hata"
            result["mesaj"]  = f"Hata: {e}"
            return result

    # In-memory bloğu dışında — arka planda kaydet (UI/kamera thread'i etkilemez)
    _save_async()
    return result


def get_today_log() -> list:
    """Bugünkü tüm hareketleri döner."""
    today = datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(JSON_LOG_PATH):
        return []
    try:
        with open(JSON_LOG_PATH, "r", encoding="utf-8") as f:
            return json.load(f).get(today, [])
    except Exception:
        return []
